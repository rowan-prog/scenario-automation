#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""CD1 셀링 포인트 시트 검증기 (v2).

규격 = config/40_selling_point_standard.md
행 번호를 고정하지 않는다 — 남색 밴드를 찾아 섹션을 잡고 그 아래를 읽는다.

사용법:
    python tools/cd1_validate.py sheet.xlsx script.md --tab "One Night with the Dragon Lord"
    (선택) --receipts read_receipts.json     # G2 미독 구간 검사

script.md = 대본을 "# Episode N" 헤더로 분할한 마크다운.
docx 에서 뽑을 때: extract-text script.docx > script.md

종료 코드 0 = 전체 통과, 1 = 실패 항목 있음.
"""
import argparse
import io
import json
import re
import sys

from openpyxl import load_workbook

# ---------------------------------------------------------------- 규격 상수

NAVY = "FF1F3864"
GRAY_HEAD = "FFBFBFBF"
LABEL_GRAY = "FFF2F2F2"
TIER_FILL = {3: "FFEA9999", 2: "FFF9CB9C", 1: "FFFFE599"}
NO_FILL = ("00000000", None)
COL_WIDTH = {"A": 33.13, "B": 35.13, "C": 38.13, "D": 42.38, "H": 51.5}

EP_ROW_START = 4
STAR_RATIO_MAX = 0.40
SCENE_MAX_CHARS = 130
KEYWORD_MIN, KEYWORD_MAX = 2, 4
TREATMENT_MIN_SENT, TREATMENT_MAX_SENT = 2, 4
MKT_BLOCK_MIN, MKT_BLOCK_MAX = 3, 5

BANNED_VERBS = {
    "가라앉": "사라진다", "되받아친": "똑같이 돌려준다", "몰아붙": "다가온다",
    "못박": "분명히 말한다", "쏘아붙": "말한다", "들이닥": "들어온다",
    "박차고": "문을 열고", "얼어붙": "굳는다", "옭아매": "붙잡아 두려고",
    "다독": "안심시킨다", "캐묻": "묻는다", "튕겨": "뒤로 날아간다",
    "소멸시": "태워 없앤다", "휩싸": "빛에 싸여", "빨려": "들어간다",
    "흘러든": "들어간다", "한 치": "밀리지 않고", "사경": "죽어감",
    "강림": "나타난", "무릅쓰": "알면서도", "짓밟": "발로 밟으며",
    "치켜": "들어 올린다", "훑": "본다", "응징": "처단",
    "폭주": "저주에 삼켜짐 / 자기를 못 알아봄",
    "밀려나": "사라진다", "내몰": "살아왔다", "끼어들": "들려온다",
    "번지": "흐른다", "굽신": "-", "핵심을": "-", "정면으로": "-",
    "무산": "막힘", "격상": "-", "종용": "시킴", "직감": "안다",
}

BANNED_EUPHEMISMS = {
    "손이 닿는 자리": "그녀가 만지는 곳마다",
    "몸을 맞댄": "젖은 실크가 허리와 다리에 달라붙은 채",
    "주도권을 놓지": "허리를 잡는 손을 떼어 침대에 붙여 놓는다",
    "밤을 주도": "침대로 밀어 넘어뜨린 뒤 위에 올라앉는다",
    "자세와 속도": "대본에 있는 동작 그대로",
    "옷 속으로 손": "치맛단 아래로 손을 넣는다",
    "마지막 밤의 자유": "새벽까지 자기 몸은 자기 것이라는 조건",
    "애무가 깊어": "대본에 있는 동작 그대로",
    "제압당한": "발로 차고 머리채를 잡는다",
    "인티머시": "-", "콜드오픈": "-", "수위 있는": "-",
}

# 템플릿 예시문 중 **서술 칸**의 것 — 여기는 English 를 써야 한다 (표준 §3-1).
# 짧은 스펙 칸(타이틀·드라이브·제작형태·언어·화수)의 예시문은 그대로 두는 게 규격이라 검사하지 않는다.
TEMPLATE_GUIDE_MARKERS = [
    "레퍼작과 무엇이 다른가를 한 문장으로",
    "선택받고 싶은 욕망 + 무시한 자들에 대한 역전",
    "여주가 스스로 판을 뒤집는다",
    "원작 있을시 원작에 대한 정보 기입",
    "여성향, 25-44",
]

EP_LIST_PATTERN = re.compile(r"EP[.\s]?\d+\s*[·,]\s*(?:EP[.\s]?)?\d+")
HANJA = re.compile(r"[一-鿿]")
EN_QUOTE = re.compile(r'"([A-Za-z][^"]{8,})"')
FAKE_AI = "AI 신규 생성 페이크"
FAKE_REUSE = "기존 장면 활용 페이크"


class Report:
    def __init__(self):
        self.rows = []

    def add(self, gate, ok, detail=""):
        self.rows.append((gate, ok, detail))

    def failed(self):
        return [r for r in self.rows if not r[1]]

    def render(self):
        out = []
        for gate, ok, detail in self.rows:
            mark = "PASS" if ok else "FAIL"
            out.append("[{0}] {1}".format(mark, gate) + ("  — " + detail if detail else ""))
        n_fail = len(self.failed())
        out.append("")
        out.append("{0}/{1} 통과".format(len(self.rows) - n_fail, len(self.rows))
                   + ("" if n_fail == 0 else ", {0}건 실패".format(n_fail)))
        return "\n".join(out)


# ---------------------------------------------------------------- 대본 파싱

def parse_script(path):
    with open(path, encoding="utf-8") as f:
        lines = f.readlines()

    starts = {}
    for i, line in enumerate(lines, 1):
        m = re.match(r"^#\s*Episode\s+(\d+)", line)
        if m:
            starts[int(m.group(1))] = i

    eps = sorted(starts)
    out = {}
    for idx, ep in enumerate(eps):
        lo = starts[ep]
        hi = starts[eps[idx + 1]] - 1 if idx + 1 < len(eps) else len(lines)
        text = "".join(lines[lo - 1:hi])
        chars = set()
        for blob in re.findall(r"\*Characters:\s*([^*]+)\*", text):
            for name in blob.split(","):
                name = re.sub(r"\s*\(.*?\)", "", name.strip().rstrip("*").strip()).strip().lower()
                if name:
                    chars.add(name)
        out[ep] = {"lo": lo, "hi": hi, "text": text, "chars": chars}
    return out, lines


# ---------------------------------------------------------------- 시트 구조 파악

def fill_of(cell):
    if cell.fill and cell.fill.patternType and cell.fill.fgColor.type == "rgb":
        rgb = cell.fill.fgColor.rgb
        return None if rgb in NO_FILL else rgb
    return None


def find_bands(ws):
    """[(row, text)] — A열 남색 밴드 전부."""
    bands = []
    for r in range(1, ws.max_row + 1):
        c = ws["A" + str(r)]
        if fill_of(c) == NAVY:
            bands.append((r, str(c.value) if c.value is not None else ""))
    return bands


def collect_episodes(ws):
    cards = {}
    r = EP_ROW_START
    while ws["F" + str(r)].value is not None:
        g = ws["G" + str(r)]
        cards[ws["F" + str(r)].value] = {
            "row": r,
            "g": g.value or "",
            "h": ws["H" + str(r)].value or "",
            "g_fill": fill_of(g),
            "f_fill": fill_of(ws["F" + str(r)]),
        }
        r += 1
    return cards


def collect_blocks(ws, bands):
    """MKT Idea / Fake 블록 = (밴드행, 헤더텍스트, [데이터행])."""
    blocks = []
    band_rows = [b[0] for b in bands]
    for i, (row, text) in enumerate(bands):
        head = text or ""
        kind = None
        if head.startswith("MKT Idea") or head.startswith("MKT Selling Point"):
            kind = "mkt"
        elif head.startswith(FAKE_AI):
            kind = "ai"
        elif head.startswith(FAKE_REUSE):
            kind = "reuse"
        elif "Fake MKT Idea" in head:
            kind = "fake_generic"
        if not kind:
            continue
        end = band_rows[i + 1] if i + 1 < len(band_rows) else ws.max_row + 1
        rows = []
        for r in range(row + 2, end):          # +1 = 회색 소제목
            if ws["A" + str(r)].value is None and ws["B" + str(r)].value is None:
                continue
            rows.append(r)
        blocks.append({"row": row, "head": head, "kind": kind, "rows": rows})
    return blocks


def tier_of(text):
    if text.startswith("★★★"):
        return 3
    if text.startswith("★★"):
        return 2
    if text.startswith("★"):
        return 1
    return 0


# ---------------------------------------------------------------- 게이트

def gate_continuity(cards, rep):
    eps = [int(e) for e in cards if str(e).strip().lstrip("EP").strip().isdigit() or isinstance(e, (int, float))]
    eps = sorted(int(e) for e in eps)
    ok = bool(eps) and eps == list(range(1, len(eps) + 1))
    rep.add("G1 회차 번호 연속성", ok,
            "" if ok else "기대 1..{0}, 실제 {1}…".format(len(eps), eps[:5]))


def gate_receipts(path, script, rep):
    if not path:
        rep.add("G2 미독 구간", True, "read_receipts.json 미제출 — 미검사")
        return
    with open(path, encoding="utf-8") as f:
        rec = json.load(f)
    bad = []
    for ep, info in script.items():
        got = rec.get(str(ep)) or rec.get(ep)
        if not got:
            bad.append("EP{0}: 기록 없음".format(ep))
            continue
        lo, hi = got.get("lo"), got.get("hi")
        if lo is None or hi is None or lo > info["lo"] or hi < info["hi"]:
            bad.append("EP{0}: {1}-{2} ⊄ {3}-{4}".format(ep, lo, hi, info["lo"], info["hi"]))
    rep.add("G2 미독 구간", not bad, "; ".join(bad[:5]))


def gate_quotes(cards, blocks, ws, script, lines, rep):
    bad = []

    def check(ep, text, where):
        if ep not in script:
            return
        lo, hi = script[ep]["lo"], script[ep]["hi"]
        for q in EN_QUOTE.findall(text or ""):
            key = q.split("...")[0].strip().rstrip(".").strip()[:30].lower()
            if not key:
                continue
            hits = [i + 1 for i, l in enumerate(lines) if key in l.lower()]
            if not any(lo <= h <= hi for h in hits):
                bad.append('{0}: "{1}" (range={2}-{3})'.format(where, q[:40], lo, hi))

    for ep, c in cards.items():
        try:
            check(int(ep), c["g"], "EP{0}".format(ep))
        except (TypeError, ValueError):
            continue
    for blk in blocks:
        for r in blk["rows"]:
            m = re.search(r"EP\s*\.?\s*(\d+)", str(ws["A" + str(r)].value or ""))
            if m:
                check(int(m.group(1)), str(ws["B" + str(r)].value or ""), "A{0}".format(r))
    rep.add("G3 인용 대사 출처", not bad, "; ".join(bad[:5]))


def gate_characters(cards, script, rep, aliases):
    if not aliases:
        rep.add("G4 인물 출처", True, "--alias 미제출 — 미검사")
        return
    bad = []
    for ep, c in cards.items():
        try:
            ep_i = int(ep)
        except (TypeError, ValueError):
            continue
        if ep_i not in script:
            continue
        blob = script[ep_i]["text"].lower()
        text = c["g"] + " " + c["h"]
        for kr, en in aliases.items():
            if en is None:          # 주연 = 회차별 호칭으로 안 바꾼다 (표준 §3-3)
                continue
            if kr in text and en.lower() not in blob:
                bad.append("EP{0}: '{1}' 언급, 대본에 '{2}' 없음".format(ep, kr, en))
    rep.add("G4 인물 출처", not bad, "; ".join(bad[:5]))


def all_cells(ws):
    for r in range(1, ws.max_row + 1):
        for col in "ABCDFGH":
            c = ws[col + str(r)]
            if isinstance(c.value, str):
                yield col + str(r), c.value


def gate_language(ws, rep):
    cells = list(all_cells(ws))

    def scan(table):
        hits = []
        for loc, text in cells:
            for bad, fix in table.items():
                if bad in text:
                    hits.append("{0}: '{1}' → {2}".format(loc, bad, fix))
        return hits

    v = scan(BANNED_VERBS)
    rep.add("G5 금지 동사·관용어", not v, "; ".join(v[:6]))
    e = scan(BANNED_EUPHEMISMS)
    rep.add("G6 완곡어법", not e, "; ".join(e[:6]))

    hanja = [loc for loc, t in cells
             if not loc.startswith("D") and HANJA.search(t) and "角色特点" not in t and "推荐场景" not in t]
    rep.add("G7 한국어 칸 한자", not hanja, ", ".join(hanja[:8]))


def gate_structure(cards, rep):
    bad = []
    for ep, c in cards.items():
        g = c["g"]
        t = tier_of(g)
        if "[키워드]" not in g:
            bad.append("EP{0}: [키워드] 없음".format(ep))
        has_scene = "[장면]" in g
        if (t > 0) != has_scene:
            bad.append("EP{0}: 별{1}인데 [장면] {2}".format(ep, t, "있음" if has_scene else "없음"))
        body = g[len("★" * t):] if t else g
        if "★" in body:
            bad.append("EP{0}: 본문에 중복 별표".format(ep))
    rep.add("G8 [키워드]/[장면] 구조", not bad, "; ".join(bad[:6]))


def gate_length(cards, rep):
    bad = []
    for ep, c in cards.items():
        g = c["g"]
        m = re.search(r"\[키워드\]\s*(.+)", g)
        if m:
            n = len([x for x in m.group(1).split("·") if x.strip()])
            if not KEYWORD_MIN <= n <= KEYWORD_MAX:
                bad.append("EP{0}: 키워드 {1}개".format(ep, n))
        m = re.search(r"\[장면\]\s*(.+)", g, re.S)
        if m and len(m.group(1).strip()) > SCENE_MAX_CHARS:
            bad.append("EP{0}: [장면] {1}자".format(ep, len(m.group(1).strip())))
        h = c["h"]
        if h:
            n_sent = len([s for s in re.split(r"(?<=다)\.\s*", h) if s.strip()])
            if not TREATMENT_MIN_SENT <= n_sent <= TREATMENT_MAX_SENT:
                bad.append("EP{0}: 트리트먼트 {1}문장".format(ep, n_sent))
    rep.add("G9 길이 규격", not bad, "; ".join(bad[:6]))


def gate_star_ratio(cards, rep):
    tiers = [tier_of(c["g"]) for c in cards.values()]
    starred = sum(1 for t in tiers if t)
    ratio = starred / len(tiers) if tiers else 0
    dist = {t: tiers.count(t) for t in (3, 2, 1) if tiers.count(t)}
    rep.add("G10 별 비율 ≤ 40%", ratio <= STAR_RATIO_MAX,
            "{0}/{1} = {2:.1%}, 분포 {3}".format(starred, len(tiers), ratio, dist))


def gate_overview(ws, bands, rep):
    start = next((r for r, t in bands if t.strip() == "Title Overview"), None)
    if start is None:
        rep.add("G11 Overview 회차 나열 없음", False, "Title Overview 밴드 없음")
        return
    bad = []
    for r in range(start + 1, start + 17):
        v = ws["B" + str(r)].value
        if isinstance(v, str) and EP_LIST_PATTERN.search(v):
            bad.append("B{0}: 회차 번호 나열".format(r))
    rep.add("G11 Overview 회차 나열 없음", not bad, "; ".join(bad))


def gate_cast(ws, bands, rep):
    start = next((r for r, t in bands if t.strip() == "CAST/CHARACTER"), None)
    if start is None:
        rep.add("G12 CAST 빌런 몰락 회차", False, "CAST 밴드 없음")
        return
    end = next((r for r, _ in bands if r > start), ws.max_row)
    bad = []
    for r in range(start + 2, end):
        name = ws["A" + str(r)].value
        body = ws["B" + str(r)].value
        if not name or not body:
            continue
        if re.search(r"villain|빌런", str(name), re.I) and not re.search(r"EP[.\s]?\d+", str(body)):
            bad.append("A{0}: 빌런 몰락 회차 미표기".format(r))
    rep.add("G12 CAST 빌런 몰락 회차", not bad, "; ".join(bad))


def gate_mkt(ws, blocks, script, rep):
    mkt = [b for b in blocks if b["kind"] == "mkt"]
    bad = []
    if not MKT_BLOCK_MIN <= len(mkt) <= MKT_BLOCK_MAX:
        bad.append("MKT Idea 블록 {0}개 (규격 {1}~{2})".format(len(mkt), MKT_BLOCK_MIN, MKT_BLOCK_MAX))
    for b in mkt:
        if not b["rows"]:
            bad.append("A{0}: 데이터 행 0".format(b["row"]))
        if b["head"].rstrip().endswith("-") or b["head"].rstrip().endswith("핵심 키워드"):
            bad.append("A{0}: 축 이름 미기입".format(b["row"]))
        for r in b["rows"]:
            m = re.search(r"EP\s*\.?\s*(\d+)", str(ws["A" + str(r)].value or ""))
            if not m:
                bad.append("A{0}: EP 표기 없음".format(r))
            elif int(m.group(1)) not in script:
                bad.append("A{0}: EP{1} 대본에 없음".format(r, m.group(1)))
    rep.add("G13 MKT Idea 블록", not bad, "; ".join(bad[:6]))


def gate_fake(ws, blocks, script, rep):
    ai = [b for b in blocks if b["kind"] == "ai"]
    reuse = [b for b in blocks if b["kind"] == "reuse"]
    generic = [b for b in blocks if b["kind"] == "fake_generic"]

    kind_bad = ["A{0}: 종류 표기 없음 (AI 신규 생성 / 기존 장면 활용)".format(b["row"]) for b in generic]
    if not ai:
        kind_bad.append("AI 신규 생성 페이크 0개")
    if not reuse:
        kind_bad.append("기존 장면 활용 페이크 0개")
    rep.add("G14a Fake 종류 표기", not kind_bad, "; ".join(kind_bad[:5]))

    reuse_bad = []
    for b in reuse:
        if "실제 서사:" not in b["head"]:
            reuse_bad.append("A{0}: '실제 서사:' 없음".format(b["row"]))
        for r in b["rows"]:
            lab = str(ws["A" + str(r)].value or "")
            m = re.search(r"EP\s*\.?\s*(\d+)", lab)
            if not m:
                reuse_bad.append("A{0}: 출처 회차 없음".format(r))
            elif int(m.group(1)) not in script:
                reuse_bad.append("A{0}: EP{1} 대본에 없음".format(r, m.group(1)))
    rep.add("G14b 기존 장면 활용 출처", not reuse_bad, "; ".join(reuse_bad[:5]))

    ai_bad = []
    for b in ai:
        if not re.search(r"\[[^\]]*\d+\s*초", b["head"]):
            ai_bad.append("A{0}: 총 길이 미표기".format(b["row"]))
        for r in b["rows"]:
            lab = str(ws["A" + str(r)].value or "")
            if not re.search(r"\d+\s*-\s*\d+\s*초", lab):
                ai_bad.append("A{0}: 컷 초 배분 없음".format(r))
    rep.add("G15 AI 신규 생성 길이", not ai_bad, "; ".join(ai_bad[:5]))


def gate_fill(cards, rep):
    bad = []
    for ep, c in cards.items():
        t = tier_of(c["g"])
        want = TIER_FILL.get(t)
        if c["f_fill"] != want or c["g_fill"] != want:
            bad.append("EP{0}: 별{1} 기대={2} F={3} G={4}".format(ep, t, want, c["f_fill"], c["g_fill"]))
    rep.add("G16 등급 ↔ 색상 일치", not bad, "; ".join(bad[:6]))


def gate_format(ws, bands, cards, rep):
    bad = []

    for row, text in bands:
        c = ws["A" + str(row)]
        f = c.font
        if (f.sz or 10) != 15 or not f.b:
            bad.append("A{0}: 밴드 글꼴 {1}pt bold={2} (기대 15/True)".format(row, f.sz or 10, f.b))
        if not (f.color and f.color.type == "rgb" and f.color.rgb.upper().endswith("FFFFFF")):
            bad.append("A{0}: 밴드 글자색 흰색 아님".format(row))

    for row, _ in bands:
        r = row + 1
        c = ws["A" + str(r)]
        if fill_of(c) == GRAY_HEAD:
            if not c.font.b or c.alignment.horizontal != "center":
                bad.append("A{0}: 소제목 bold/가운데 아님".format(r))

    for loc, text in all_cells(ws):
        for marker in TEMPLATE_GUIDE_MARKERS:
            if marker in text:
                bad.append("{0}: 서술 칸에 예시문 잔존 (English 를 써야 한다)".format(loc))

    for r in range(1, ws.max_row + 1):
        for col in "ABCDFGH":
            c = ws[col + str(r)]
            if c.value is None:
                continue
            if c.alignment.vertical != "top" or not c.alignment.wrap_text:
                bad.append("{0}{1}: top/wrap 아님".format(col, r))

    for col, want in COL_WIDTH.items():
        got = ws.column_dimensions[col].width if col in ws.column_dimensions else None
        if got is None or abs(got - want) > 0.5:
            bad.append("{0}열 너비 {1} (기대 {2})".format(col, got, want))
    g_width = ws.column_dimensions["G"].width if "G" in ws.column_dimensions else None
    if cards and any("[장면]" in c["g"] for c in cards.values()) and (g_width or 0) < 40:
        bad.append("G열 너비 {0} — [장면]을 쓰면 45.5".format(g_width))

    rep.add("G17 서식 규격", not bad, "; ".join(bad[:8]) + (" …외 {0}건".format(len(bad) - 8) if len(bad) > 8 else ""))


# ---------------------------------------------------------------- 진입점

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("script")
    ap.add_argument("--tab", required=True)
    ap.add_argument("--receipts", help="read_receipts.json (G2)")
    ap.add_argument("--alias", help='{"클로이": "Chloe", "이솔데": null} 형태 json (G4). '
                                    'null = 주연 — 그 회차에 이름이 안 나와도 이름으로 부르므로 검사 제외')
    ap.add_argument("--json", help="결과 JSON 저장 경로")
    args = ap.parse_args()

    script, lines = parse_script(args.script)
    wb = load_workbook(args.xlsx)
    if args.tab not in wb.sheetnames:
        print("탭 없음: {0} (있는 탭: {1})".format(args.tab, wb.sheetnames))
        return 1
    ws = wb[args.tab]

    bands = find_bands(ws)
    cards = collect_episodes(ws)
    blocks = collect_blocks(ws, bands)
    aliases = json.load(open(args.alias, encoding="utf-8")) if args.alias else None

    rep = Report()
    gate_continuity(cards, rep)
    gate_receipts(args.receipts, script, rep)
    gate_quotes(cards, blocks, ws, script, lines, rep)
    gate_characters(cards, script, rep, aliases)
    gate_language(ws, rep)
    gate_structure(cards, rep)
    gate_length(cards, rep)
    gate_star_ratio(cards, rep)
    gate_overview(ws, bands, rep)
    gate_cast(ws, bands, rep)
    gate_mkt(ws, blocks, script, rep)
    gate_fake(ws, blocks, script, rep)
    gate_fill(cards, rep)
    gate_format(ws, bands, cards, rep)

    print(rep.render())
    print("")
    print("집계: 회차 {0} · ★★★ {1} · ★★ {2} · ★ {3} · MKT 블록 {4} · Fake 블록 {5}".format(
        len(cards),
        sum(1 for c in cards.values() if tier_of(c["g"]) == 3),
        sum(1 for c in cards.values() if tier_of(c["g"]) == 2),
        sum(1 for c in cards.values() if tier_of(c["g"]) == 1),
        sum(1 for b in blocks if b["kind"] == "mkt"),
        sum(1 for b in blocks if b["kind"] in ("ai", "reuse", "fake_generic"))))

    if args.json:
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump([{"gate": g, "ok": o, "detail": d} for g, o, d in rep.rows],
                      f, ensure_ascii=False, indent=1)
    return 1 if rep.failed() else 0


if __name__ == "__main__":
    sys.exit(main())
