#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""CD1 셀링 포인트 시트 빌더.

spec.json 을 받아 [EN_AI] Selling Point 템플릿 규격 그대로 탭 하나를 만든다.
서식(색·글자색·사이즈·정렬·테두리·열 너비)은 전부 여기서 명시적으로 칠한다.
템플릿에서 상속된 서식에 기대지 않는다 — 남은 fill 때문에 글자가 사라지는 사고 방지.

사용법:
    python tools/build_cd1_sheet.py spec.json --template "<템플릿.xlsx>" -o out.xlsx
    python tools/build_cd1_sheet.py --example > spec.json

규격 = config/40_selling_point_standard.md §2
"""
import argparse
import json
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------- 서식 상수 (템플릿 실측)

FONT_NAME = "Arial"
NAVY = "FF1F3864"
GRAY_HEAD = "FFBFBFBF"
LABEL_GRAY = "FFF2F2F2"
BORDER_RGB = "FFC9C9C9"

TIER_FILL = {3: "FFEA9999", 2: "FFF9CB9C", 1: "FFFFE599"}
TIER_STAR = {3: "★★★", 2: "★★", 1: "★"}

COL_WIDTH = {"A": 33.13, "B": 35.13, "C": 38.13, "D": 42.38, "G": 16.5, "H": 51.5}
# G=16.5 는 템플릿 원본값. 넓히면 팀 워크북에 붙였을 때 우리 탭만 어긋난다 (표준 §2-3).

STATUS = {
    "not_started": "⚪ 작성대기 / Not Started",
    "in_progress": "🟡 작성중 / In Progress",
    "complete": "🟢 작성완료 / Complete",
}

THIN = Side(style="thin", color=BORDER_RGB)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_BORDER = Border()
NO_FILL = PatternFill(fill_type=None)

TOP_LEFT = Alignment(horizontal=None, vertical="top", wrap_text=True)
TOP_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
BAND_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

OVERVIEW_ROWS = [
    ("title", "타이틀 / Title"),
    ("program_id", "프로그램 코드 / Program ID "),
    ("pm", "기획 담당 / Project Manager"),
    ("drive", "Project Google Drive"),
    ("release", "론칭일 / Release Date "),
    ("format", "제작 형태 / Format"),
    ("language", "발화 언어 / Language "),
    ("eps", "Total EP / Paywall EP"),
    ("logline", "Logline"),
    ("synopsis", "Synopsis"),
    ("genre", "장르 키워드 (3개) / Genre Keywords (3)"),
    ("differentiator", "핵심 차별점 (1문장) / Key Differentiator (1 sentence)"),
    ("emotion", "타겟이 반응할 감정·욕망 / Target Emotion & Desire"),
    ("highlight", "반드시 강조해야 할 요소 / Must-Highlight Element"),
    ("target", "기획 타겟 정보 / Target Audience"),
    ("other", "기타 주요 정보 / Other Key Info"),
]

REFERENCE_LABELS = [
    ("title", "레퍼런스 작품 / Reference Title"),
    ("scenes", "레퍼작의 핵심 장면 및 내용 / Key Scenes & Content of Reference"),
    ("ad", "광고소재 (확인 가능 시) / Ad Creative (if available)"),
]
ORIGINAL_IP_LABELS = [
    ("title", "원작 IP / Original IP "),
    ("scenes", "원작 IP의 핵심 장면 및 내용 / Key Scenes & Content of Original IP "),
    ("ad", "광고소재 (확인 가능 시) / Ad Creative (if available)"),
]

EPISODIC_HEAD = [
    ("F", "EP"),
    ("G", "주요 사건 키워드 / Key Event Keywords"),
    ("H", "스토리 요약 (2-3문장) / Story Summary (2-3 sentences)"),
]
SCENE_HEAD = ["추천 장면 (KR)", "Recommended Scene (EN)", "\t推荐场景 (CN)"]
MKT_HEAD_PREFIX = "MKT Idea (Copy this format if there's more than 1) - "

# C열 예시문 — spec 의 en 이 null 이면 이 문구를 회색 9pt 이탤릭 그대로 남긴다
#  (짧은 스펙 칸은 B(한국어)에 값이 들어가면 C의 영어가 중복이라 예시문을 유지한다 — 표준 §3-1)
GUIDE_TEXT = {
    "title": "(eg) Bound to the Alpha",
    "drive": "video, srt, script 모두 이 안에서 확인 / Video, SRT and script files ",
    "format": "AI 실사 / AI Live Action",
    "language": "English / USA",
    "eps": "72화 / 페이월 8화",
    "differentiator": "레퍼작과 무엇이 다른가를 한 문장으로 / What makes this different from the reference title, in one line",
    "emotion": "선택받고 싶은 욕망 + 무시한 자들에 대한 역전 / Desire to be chosen + payback against those who dismissed her",
    "highlight": "여주가 스스로 판을 뒤집는다 (구조받는 서사 아님) / The FL flips the table herself (not a rescued heroine)",
    "target": "여성향, 25-44 / Female-skewing, 25-44",
    "other": "원작 있을시 원작에 대한 정보 기입 / If adapted from existing IP, note the original here",
}
GUIDE_RGB = "FF808080"

MIN_CAST_ROWS = 9          # 기본 레이아웃(Reference 밴드 = 32행)을 지키는 최소치
SCENE_MAX_CHARS = 130


# ---------------------------------------------------------------- 셀 유틸

def wipe(ws, row, cols="ABCDEFGH"):
    """상속 서식 제거 — 값·채우기·테두리·글꼴 초기화."""
    for col in cols:
        c = ws[col + str(row)]
        c.value = None
        c.fill = NO_FILL
        c.border = NO_BORDER
        c.font = Font(name=FONT_NAME)
        c.alignment = TOP_LEFT


def band(ws, row, cols, text, merge=True):
    """남색 섹션 밴드. Arial 15 Bold 흰색."""
    for col in cols:
        c = ws[col + str(row)]
        c.value = None
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name=FONT_NAME, size=15, bold=True, color="FFFFFFFF")
        c.alignment = BAND_ALIGN
        c.border = NO_BORDER
    if merge and len(cols) > 1:
        ws.merge_cells("{0}{2}:{1}{2}".format(cols[0], cols[-1], row))
    ws[cols[0] + str(row)] = text


def subhead(ws, row, values, start_col="A"):
    """회색 소제목 행. Bold·가운데·테두리."""
    base = ord(start_col)
    for i, v in enumerate(values):
        c = ws[chr(base + i) + str(row)]
        c.value = v
        c.fill = PatternFill("solid", fgColor=GRAY_HEAD)
        c.font = Font(name=FONT_NAME, bold=True)
        c.alignment = TOP_CENTER
        c.border = BOX


def label(ws, row, text, col="A"):
    c = ws[col + str(row)]
    c.value = text
    c.fill = PatternFill("solid", fgColor=LABEL_GRAY)
    c.font = Font(name=FONT_NAME, bold=True)
    c.alignment = TOP_LEFT
    c.border = BOX


def data(ws, row, col, value, center=False, fill=None, bold=False, guide=False):
    c = ws[col + str(row)]
    c.value = value
    c.fill = PatternFill("solid", fgColor=fill) if fill else NO_FILL
    c.font = (Font(name=FONT_NAME, size=9, italic=True, color=GUIDE_RGB) if guide
              else Font(name=FONT_NAME, bold=bold))
    c.alignment = TOP_CENTER if center else TOP_LEFT
    c.border = BOX


# ---------------------------------------------------------------- 섹션 빌드

def build_overview(ws, spec):
    band(ws, 1, "ABC", spec.get("title") or spec["tab"])
    st = spec.get("status", "complete")
    if st not in STATUS:
        raise SystemExit("status 는 " + " / ".join(STATUS) + " — 받은 값: " + repr(st))
    d1 = ws["D1"]
    d1.value = STATUS[st]
    d1.fill = NO_FILL
    d1.font = Font(name=FONT_NAME, size=11, bold=True)
    d1.alignment = TOP_LEFT
    d1.border = NO_BORDER
    band(ws, 3, "ABC", "Title Overview", merge=False)
    ov = spec.get("overview") or {}
    for i, (key, text) in enumerate(OVERVIEW_ROWS):
        r = 4 + i
        label(ws, r, text)
        item = ov.get(key) or {}
        if isinstance(item, str):
            item = {"kr": item, "en": ""}
        data(ws, r, "B", item.get("kr") or None)
        en = item.get("en")
        if en is None:                      # 예시문 유지 (회색 9pt 이탤릭)
            data(ws, r, "C", GUIDE_TEXT.get(key), guide=True)
        else:
            data(ws, r, "C", en or None)
    wipe(ws, 20, "ABCD")


def build_episodic(ws, spec):
    band(ws, 2, "FGH", "Episodic Breakdown", merge=False)
    for col, text in EPISODIC_HEAD:
        c = ws[col + "3"]
        c.value = text
        c.fill = PatternFill("solid", fgColor=GRAY_HEAD)
        c.font = Font(name=FONT_NAME, bold=True)
        c.alignment = TOP_CENTER
        c.border = BOX

    eps = spec.get("episodes") or []
    warn = []
    for i, ep in enumerate(eps):
        r = 4 + i
        tier = int(ep.get("tier") or 0)
        fill = TIER_FILL.get(tier)
        kws = ep.get("keywords") or []
        if not 2 <= len(kws) <= 4:
            warn.append("EP{0}: 키워드 {1}개 (2~4)".format(ep.get("ep"), len(kws)))
        lines = []
        if tier:
            lines.append(TIER_STAR[tier])
        lines.append("[키워드] " + " · ".join(kws))
        scene = (ep.get("scene") or "").strip()
        if tier and not scene:
            warn.append("EP{0}: 별{1}인데 [장면] 없음".format(ep.get("ep"), tier))
        if scene and not tier:
            warn.append("EP{0}: 무별점인데 [장면] 있음".format(ep.get("ep")))
        if scene:
            if len(scene) > SCENE_MAX_CHARS:
                warn.append("EP{0}: [장면] {1}자 (>{2})".format(ep.get("ep"), len(scene), SCENE_MAX_CHARS))
            lines.append("[장면] " + scene)
        data(ws, r, "F", ep.get("ep"), center=True, fill=fill)
        data(ws, r, "G", "\n".join(lines), fill=fill)
        data(ws, r, "H", ep.get("treatment"))

    # 템플릿에 미리 그려진 남는 행 정리
    for r in range(4 + len(eps), 4 + max(len(eps), 60) + 40):
        for col in "FGH":
            c = ws[col + str(r)]
            c.value = None
            c.fill = NO_FILL
            c.border = NO_BORDER
    return warn


def build_cast(ws, spec, start=21):
    band(ws, start, "ABCD", "CAST/CHARACTER", merge=False)   # 템플릿 원본은 병합 없이 네 칸 각각 칠함
    subhead(ws, start + 1, ["캐릭터명 / Character Name", "캐릭터 정보", "characteristic", "角色特点"])
    cast = spec.get("cast") or []
    rows = max(len(cast), MIN_CAST_ROWS)
    for i in range(rows):
        r = start + 2 + i
        person = cast[i] if i < len(cast) else {}
        data(ws, r, "A", person.get("name"), bold=bool(person.get("name")))
        data(ws, r, "B", person.get("kr"))
        data(ws, r, "C", person.get("en"))
        data(ws, r, "D", person.get("cn"))
    return start + 2 + rows          # 다음 섹션 시작 행


def build_reference(ws, spec, start):
    is_ip = spec.get("reference_label") == "Original IP"
    labels = ORIGINAL_IP_LABELS if is_ip else REFERENCE_LABELS
    title = "Original IP" if is_ip else "Reference"
    refs = spec.get("reference") or [{}]
    r = start
    for ref in refs:
        band(ws, r, "ABCD", title, merge=False)
        r += 1
        for key, text in labels:
            label(ws, r, text)
            data(ws, r, "B", ref.get(key))
            data(ws, r, "C", ref.get(key + "_en"))
            data(ws, r, "D", ref.get(key + "_cn"))
            r += 1
    return r


def build_blocks(ws, spec, start):
    """MKT Idea 블록 → Fake MKT Idea 블록."""
    r = start
    warn = []
    for blk in spec.get("mkt_ideas") or []:
        # head 를 주면 밴드 문구를 통째로 갈아 끼운다(모범 탭 방식).
        # 안 주면 템플릿 원문 + 축 이름.
        band(ws, r, "ABCD", blk.get("head") or (MKT_HEAD_PREFIX + blk.get("axis", "")))
        subhead(ws, r + 1, ["EP"] + SCENE_HEAD)
        r += 2
        rows = blk.get("rows") or []
        noline = 0
        for row in rows:
            kr = row.get("kr") or ""
            if '"' not in kr and "“" not in kr and "「" not in kr:
                noline += 1
            data(ws, r, "A", row.get("ep"), center=True)
            data(ws, r, "B", kr or None)
            data(ws, r, "C", row.get("en"))
            data(ws, r, "D", row.get("cn"))
            r += 1
        if rows and noline * 2 > len(rows):
            warn.append("MKT '{0}': 대사 없는 행 {1}/{2} — 절반 넘으면 소재가 안 된다 (표준 §3-7)".format(
                blk.get("axis") or blk.get("head") or "?", noline, len(rows)))
        wipe(ws, r, "ABCD")
        r += 1

    for blk in spec.get("fakes") or []:
        kind = blk.get("kind")
        no = blk.get("no", "")
        if kind == "ai":
            head = "AI 신규 생성 페이크 {0} — {1} [{2}]".format(no, blk.get("title", ""), blk.get("spec", ""))
            if "초" not in str(blk.get("spec", "")):
                warn.append("AI 페이크 {0}: 총 길이(초) 미표기".format(no))
        elif kind == "reuse":
            head = "기존 장면 활용 페이크 {0} — {1} [{2}]".format(no, blk.get("title", ""), blk.get("spec", ""))
            truth = (blk.get("truth") or "").strip()
            if not truth:
                warn.append("기존 장면 페이크 {0}: '실제 서사:' 없음".format(no))
            else:
                if not truth.startswith("실제 서사:"):
                    truth = "실제 서사: " + truth
                head += "\n" + truth
        else:
            raise SystemExit("fakes[].kind 는 'ai' 또는 'reuse' — 받은 값: " + repr(kind))
        band(ws, r, "ABCD", head)
        subhead(ws, r + 1, ["No."] + SCENE_HEAD)
        r += 2
        for cut in blk.get("cuts") or []:
            lab = cut.get("label", "")
            if kind == "ai" and "초" not in lab:
                warn.append("AI 페이크 {0} {1}: 컷 초 배분 없음".format(no, lab))
            if kind == "reuse" and "EP" not in lab.upper():
                warn.append("기존 장면 페이크 {0} {1}: 출처 회차 없음".format(no, lab))
            data(ws, r, "A", lab, center=True)
            data(ws, r, "B", cut.get("kr"))
            data(ws, r, "C", cut.get("en"))
            data(ws, r, "D", cut.get("cn"))
            r += 1
        wipe(ws, r, "ABCD")
        r += 1
    return r, warn


# ---------------------------------------------------------------- 진입점

def build(spec, template_path, out_path):
    if template_path:
        wb = load_workbook(template_path)
        for name in list(wb.sheetnames):
            if name != "Template":
                del wb[name]
    else:
        wb = Workbook()
        wb.remove(wb.active)

    tab = spec["tab"][:31]
    if tab in wb.sheetnames:
        del wb[tab]
    ws = wb.create_sheet(tab)

    for col, w in COL_WIDTH.items():
        ws.column_dimensions[col].width = w

    build_overview(ws, spec)
    warn = build_episodic(ws, spec)
    after_cast = build_cast(ws, spec)
    after_ref = build_reference(ws, spec, after_cast)
    wipe(ws, after_ref, "ABCD")
    wipe(ws, after_ref + 1, "ABCD")
    last, w2 = build_blocks(ws, spec, after_ref + 2)   # 템플릿 기본: 36~37 공백 후 38행
    warn += w2

    wb.save(out_path)
    return warn, last


EXAMPLE = {
    "tab": "One Night with the Dragon Lord",
    "status": "complete",
    "title": "One Night with the Dragon Lord",
    "overview": {
        "title": {"kr": "One Night with the Dragon Lord", "en": "One Night with the Dragon Lord"},
        "program_id": {"kr": "", "en": ""},
        "pm": {"kr": "", "en": ""},
        "drive": {"kr": "", "en": ""},
        "release": {"kr": "", "en": ""},
        "format": {"kr": "AI 실사", "en": "AI Live Action"},
        "language": {"kr": "EN", "en": "EN / USA"},
        "eps": {"kr": "총 53화 / 무료 1~8화 · 페이월 9화", "en": "53 / Free EP 1-8, paywall at EP9"},
        "logline": {"kr": "…", "en": "…"},
        "synopsis": {"kr": "…", "en": "…"},
        "genre": {"kr": "정략결혼 · 혐관로맨스 · 다크판타지",
                  "en": "Arranged Marriage · Enemies-to-Lovers · Dark Fantasy"},
        "differentiator": {"kr": "…", "en": "…"},
        "emotion": {"kr": "…", "en": "…"},
        "highlight": {"kr": "…", "en": "…"},
        "target": {"kr": "여성향. 20~30대. …", "en": "Female-skewing, 20s-30s. …"},
        "other": {"kr": "원작 없는 오리지널.", "en": "Original IP, no source novel."}
    },
    "episodes": [
        {"ep": 1, "tier": 3,
         "keywords": ["정사신", "여주주도", "정체반전"],
         "scene": "괴물에게 팔려가기 전날 밤, 이솔데가 남자의 입술을 물어뜯고 침대로 밀어 넘어뜨린 뒤 위에 올라앉아 셔츠를 찢는다. \"Shut up... and serve me.\"",
         "treatment": "폭풍우 치는 여관방. …"},
        {"ep": 2, "tier": 0, "keywords": ["간통누명", "위협"], "scene": None, "treatment": "…"}
    ],
    "cast": [
        {"name": "이솔데 (Isolde) / FL", "kr": "…", "en": "…", "cn": "…"}
    ],
    "reference_label": "Reference",
    "reference": [
        {"title": "…", "scenes": "…", "ad": "https://…"}
    ],
    "mkt_ideas": [
        {"axis": "첫날밤 상대를 자기가 고른다",
         "rows": [{"ep": "EP1", "kr": "…", "en": "…", "cn": "…"}]}
    ],
    "fakes": [
        {"kind": "ai", "no": 1, "title": "남자들을 줄 세워 놓고 오늘 밤 잘 상대를 고르는 여자",
         "spec": "21초 · 4컷",
         "cuts": [{"label": "컷1 · 0-5초", "kr": "△ …", "en": "△ …", "cn": "△ …"}]},
        {"kind": "reuse", "no": 1, "title": "첫날밤마다 신부를 찢어 죽인다는 남자",
         "spec": "4컷 · EP1·5·8·34",
         "truth": "바엘은 이솔데에게 폭력을 쓰지 않는다. 태워 죽이는 대상은 전부 그녀를 해치러 온 자들이다.",
         "cuts": [{"label": "컷1 · EP1", "kr": "△ …", "en": "△ …", "cn": "△ …"}]}
    ]
}


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    ap = argparse.ArgumentParser()
    ap.add_argument("spec", nargs="?", help="spec.json")
    ap.add_argument("--template", help="[EN_AI] Selling Point 템플릿 xlsx (Template 탭 보존용)")
    ap.add_argument("-o", "--out", help="출력 xlsx")
    ap.add_argument("--example", action="store_true", help="spec 스켈레톤 출력")
    args = ap.parse_args()

    if args.example:
        json.dump(EXAMPLE, sys.stdout, ensure_ascii=False, indent=2)
        print()
        return 0
    if not args.spec or not args.out:
        ap.error("spec.json 과 -o 가 필요하다 (또는 --example)")

    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)
    warn, last = build(spec, args.template, args.out)
    print("생성: {0}  탭 '{1}'  회차 {2}  캐스트 {3}  MKT {4}  Fake {5}  마지막 행 {6}".format(
        args.out, spec["tab"], len(spec.get("episodes") or []), len(spec.get("cast") or []),
        len(spec.get("mkt_ideas") or []), len(spec.get("fakes") or []), last))
    if warn:
        print("경고 (검증기 게이트에서 걸린다):")
        for w in warn:
            print("  -", w)
    return 0


if __name__ == "__main__":
    sys.exit(main())
